import openpyxl
import os
import asyncio
import smtplib
import unicodedata
from email.message import EmailMessage
from playwright.async_api import async_playwright
from Functions.UserInfo import jira_url, jira_username, jira_password, data_file_path, network_data_path, security_test_path
from Functions.read_excel_file import read_excel_file
from Steps.login import login
from Steps.response import response
from Steps.run_query import run_query
from Steps.click_chat_box import click_chat_box
import smtplib
from email.message import EmailMessage
import time


# Hàm loại bỏ dấu tiếng Việt trong chuỗi
def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

# Hàm gửi email
# Hàm gửi email qua SMTP nội bộ
import requests

def send_email_via_gmail(file_path, summary):
    # Thông tin cấu hình Gmail
    smtp_server = "smtp.gmail.com"  # SMTP server của Gmail
    port = 587  # Cổng SMTP với TLS
    sender_email = "hpt.has.hcm@gmail.com"  # Email gửi đi
    sender_password = "xsdk kqag osez ljbd"  # App Password của Gmail
    receiver_emails = ["khuongn@hpt.vn", "khuongnguyen04061999@gmail.com"]  # Email người nhận

    # Tạo nội dung email
    subject = "Kết quả kiểm thử tự động - [Dự án MVP – AI bổ sung nguồn lực]"
    body = (
        f"1) Số lượng câu hỏi đã được kiểm thử: {summary['total']}\n"
        f"2) Số lượng câu hỏi trả về từ Database: {summary['database_results']}\n"
        f"3) Số lượng câu hỏi trả về từ chatbot AI: {summary['response_results']}\n"
        f"5) Số lượng câu hỏi FAIL: {summary['fail']}\n"
        f"6) Thời gian hoàn thành bộ testcase: {summary['execution_time']} phút"
    )

    # Tạo email
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = ", ".join(receiver_emails)
    msg.set_content(body)

    # Đính kèm file kết quả
    try:
        with open(file_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=file_path.split("/")[-1],
            )

        # Gửi email qua SMTP Gmail
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls()  # Kích hoạt bảo mật TLS
            server.login(sender_email, sender_password)  # Đăng nhập vào Gmail
            server.send_message(msg)  # Gửi email

        print("Email đã được gửi thành công qua Gmail!")
    except Exception as e:
        print(f"Lỗi khi gửi email: {e}")

# Gửi email
# send_email_via_smtp("result_file.xlsx", summary)



# Hàm lưu kết quả vào file Excel
def save_to_excel(file_path, data):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for idx, record in enumerate(data, start=2):
            # Cột 4: Ghi giá trị response
            sheet.cell(row=idx, column=4, value=record.get("response", ""))

            # Cột 5: Ghi toàn bộ query_result dưới dạng chuỗi y nguyên
            query_result = record.get("query_result", "")
            if isinstance(query_result, list):
                # Ghi y nguyên toàn bộ nội dung Record trong query_result
                result_str = "\n".join([str(row) for row in query_result])
            else:
                result_str = str(query_result)
            sheet.cell(row=idx, column=5, value=result_str)

            # Cột 7: Ghi status
            sheet.cell(row=idx, column=7, value=record.get("status", "FAIL"))

        # Lưu file mới với hậu tố _result.xlsx
        new_file_path = file_path.replace(".xlsx", "_result.xlsx")
        workbook.save(new_file_path)
        print(f"Results saved to {new_file_path}")
        return new_file_path
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return None



# Hàm thống kê kết quả
def summarize_results(data):
    return {
        "total": len(data),
        "database_results": sum(1 for record in data if record.get("query_result")),
        "response_results": sum(1 for record in data if record.get("response")),
        "pass": sum(1 for record in data if record.get("status") == "PASS"),
        "fail": sum(1 for record in data if record.get("status") == "FAIL"),
    }

# Hàm xử lý workflow
async def process_workflow(playwright, data, start_index=0, run_all=False):
    os.makedirs('screenshots', exist_ok=True)  # Tạo thư mục lưu ảnh chụp màn hình nếu chưa có

    browser = await playwright.chromium.launch(headless=False)  # Mở trình duyệt
    context = await browser.new_context(ignore_https_errors=True)  # Bỏ qua lỗi chứng chỉ HTTPS
    page = await context.new_page()  # Tạo trang mới

    try:
        # Mở trang web và thực hiện đăng nhập, click vào chat box
        await page.goto("http://10.4.18.152:3001/")
        await login(page)
        await click_chat_box(page)

        index = start_index
        while index < len(data):
            record = data[index]  # Lấy dữ liệu của testcase hiện tại
            print(f"Running testcase {index + 1}")  # In ra số testcase đang chạy

            # Kiểm tra điều kiện refresh trang
            if index > 0 and record.get("IsRefresh", "").upper() == "YES":
                print(f"Refreshing page for testcase {index + 1} as IsRefresh = YES")
                await page.reload()  # Refresh trang
                await page.wait_for_load_state("networkidle")
                await click_chat_box(page)  # Chạy lại hàm click_chat_box để đảm bảo trang sẵn sàng

            try:
                # Gọi hàm lấy phản hồi từ chatbot
                latest_response = await response(page, record)
                record["response"] = latest_response

                # Chạy truy vấn query và lấy kết quả
                query_result = await run_query(record["query"])
                if isinstance(query_result, list):
                    query_result = [list(row.values())[0] for row in query_result]
                elif hasattr(query_result, 'values'):
                    query_result = list(query_result.values())[0]

                # So sánh kết quả phản hồi với dữ liệu từ query
                status = "FAIL"
                normalized_response = remove_accents(latest_response).lower().strip()
                if isinstance(query_result, list):
                    for result in query_result:
                        if remove_accents(str(result)).lower().strip() in normalized_response:
                            status = "PASS"
                            break
                elif remove_accents(str(query_result)).lower().strip() in normalized_response:
                    status = "PASS"

                # Lưu kết quả vào record
                record["query_result"] = query_result
                record["status"] = status

            except Exception as e:
                print(f"Error on testcase {index + 1}: {e}")

            # Dừng xử lý nếu không chạy tất cả các case
            if not run_all:
                choice = input(f"Run the next test case? (Yes/No): ")
                if choice.lower() != 'yes':
                    break

            index += 1  # Chuyển sang testcase tiếp theo

    finally:
        await context.close()  # Đóng context của trình duyệt
        await browser.close()  # Đóng trình duyệt


# Hàm bắt đầu workflow
async def start_workflow(playwright):
    start_time = time.time()  # Ghi lại thời gian bắt đầu
    data = read_excel_file(data_file_path)
    choice = input("Run all test cases or just 1? (type 'all' or a number): ")

    if choice.lower() == 'all':
        await process_workflow(playwright, data, run_all=True)
    else:
        try:
            specific_test = int(choice)
            await process_workflow(playwright, data, start_index=specific_test)
        except ValueError:
            print("Invalid input. Please type 'all' or a number.")

    # Lưu kết quả
    result_file = save_to_excel(data_file_path, data)
    summary = summarize_results(data)
    
    # Tính thời gian hoàn thành
    end_time = time.time()
    elapsed_time_minutes = round((end_time - start_time) / 60, 2)
    summary["execution_time"] = elapsed_time_minutes  # Thêm vào summary

    # Gửi email kết quả
    send_email_via_gmail(result_file, summary)


# Hàm main
async def main():
    async with async_playwright() as playwright:
        await start_workflow(playwright)

if __name__ == "__main__":
    asyncio.run(main())
