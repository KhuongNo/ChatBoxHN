from datetime import datetime
from playwright.async_api import Page
from openpyxl import Workbook, load_workbook
import os

async def test_xss(page: Page, url: str, output_file: str):
    errors = []

    # Event listener for console errors (e.g., script execution)
    page.on("console", lambda message: message.type == "error" and errors.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "message": message.text,
        "location": message.location if hasattr(message, 'location') else None,
        "stack": None
    }))

    # Event listener for page errors (e.g., script errors)
    page.on("pageerror", lambda error: errors.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "message": error.message,
        "name": error.name,
        "stack": error.stack
    }))

    # Example: Injecting XSS payload into a search form
    try:
        await page.goto(url)  # Use the provided URL
        await page.fill('input[name="q"]', '"><script>alert("XSS")</script>')
        await page.press('input[name="q"]', 'Enter')
        await page.wait_for_timeout(5000)  # Wait for some time to capture errors

    except Exception as e:
        errors.append({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "message": str(e),
            "name": type(e).__name__,
            "stack": None
        })

    # Load or create the workbook
    if os.path.exists(output_file):
        workbook = load_workbook(output_file)
        errors_sheet = workbook.active
    else:
        workbook = Workbook()
        errors_sheet = workbook.create_sheet("XSS Test Results", 0)
        errors_headers = ["Time", "URL", "Message", "Name", "Location/Stack"]
        errors_sheet.append(errors_headers)

    # Write errors data to the errors sheet
    if errors:
        for error in errors:
            location_or_stack = error.get("stack") or error.get("location")
            errors_sheet.append([error["time"], url, error["message"], error.get("name", ""), location_or_stack])
    else:
        errors_sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), url, "No XSS found", "", ""])

    # Save the workbook to the specified output file
    workbook.save(output_file)
